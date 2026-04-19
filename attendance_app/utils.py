import random
from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO

from openpyxl import Workbook


def today_iso():
    return date.today().isoformat()


def one_year_ago_iso():
    return (date.today() - timedelta(days=365)).isoformat()


def build_otp():
    return f"{random.randint(100000, 999999)}"


def otp_expiry(minutes):
    return (datetime.now() + timedelta(minutes=minutes)).strftime("%Y-%m-%d %H:%M:%S")


def is_valid_mobile(number):
    return number.isdigit() and len(number) == 10


def attendance_workbook(rows):
    wb = Workbook()
    headers = [
        "Date",
        "Month",
        "Institution",
        "Class Roll Number",
        "Roll Number",
        "Student Name",
        "Branch",
        "Semester",
        "Year",
        "Status",
        "Marked By",
        "Remarks",
    ]
    monthly_rows = defaultdict(list)
    daily_rows = defaultdict(list)

    for row in rows:
        month_key = row["attendance_date"][:7]
        monthly_rows[month_key].append(row)
        daily_rows[row["attendance_date"]].append(row)

    def populate_sheet(ws, sheet_rows):
        ws.append(headers)
        for row in sheet_rows:
            ws.append(
                [
                    row["attendance_date"],
                    row["attendance_date"][:7],
                    row["institution_name"] or "",
                    row["class_roll_number"] or "",
                    row["roll_number"],
                    row["full_name"],
                    row["branch"],
                    row["semester"],
                    row["year"],
                    row["status"],
                    row["marked_by_name"],
                    row["remarks"] or "",
                ]
            )

    summary = wb.active
    summary.title = "Summary"
    populate_sheet(summary, rows)

    for month_key in sorted(monthly_rows):
        ws = wb.create_sheet(title=f"Month-{month_key}"[:31])
        populate_sheet(ws, monthly_rows[month_key])

    for day_key in sorted(daily_rows):
        ws = wb.create_sheet(title=f"Day-{day_key}"[:31])
        populate_sheet(ws, daily_rows[day_key])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
